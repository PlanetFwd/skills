#!/usr/bin/env python3
"""
PlanetFWD Country of Origin Validator
======================================
Validates raw country of origin (COO) data against the official PlanetFWD
list of acceptable country names and produces an annotated output spreadsheet.

Usage:
    python3 validate_coo.py <input_file> [--col <column_name>] [--sheet <sheet_name>]
                            [--output <output_file>] [--skill-dir <path>]
                            [--mode mapping|full]

Arguments:
    input_file      Path to the input .xlsx, .csv, or .tsv file
    --col           Name of the column containing raw COO data (default: auto-detect)
    --sheet         Sheet name if input is Excel (default: first sheet)
    --output        Output file path (default: <input_file>_validated.xlsx)
    --skill-dir     Path to the country_validation skill directory (for loading assets)
    --mode          Output mode:
                      mapping (default) — one row per unique raw COO value, mapping columns only
                      full             — all original columns retained, validated columns appended

Output modes:
    mapping: Unique COO mapping table (deduplicated). Columns:
               Raw Country of Origin | First Country Parsed |
               Validated COO (PlanetFWD) | Match Method

    full:    Original dataset with three columns appended to the right:
               First Country Parsed | Validated COO (PlanetFWD) | Match Method
"""

import sys
import json
import argparse
import re
import pandas as pd
from pathlib import Path


# ─────────────────────────────────────────────
# REGIONAL / CONTINENTAL TERMS (→ always "Unknown")
# These are geographic regions, not countries, and cannot be
# mapped to any single PlanetFWD country entry.
# ─────────────────────────────────────────────
REGIONAL_TERMS = {
    "asia", "africa", "europe", "south america", "north america",
    "central america", "middle east", "oceania", "antarctica",
    "caribbean", "west africa", "east africa", "southeast asia",
    "south asia", "north africa", "sub-saharan africa", "latin america",
    "global", "worldwide", "international", "various", "multiple",
    "unknown", "other", "n/a", "na", "none", "not applicable",
    "not specified", "unspecified",
}


def load_assets(skill_dir: str) -> tuple[set, dict, dict]:
    """Load valid countries list and alias mapping from skill's assets folder."""
    assets_dir = Path(skill_dir) / "assets"

    valid_path = assets_dir / "valid_countries.json"
    if not valid_path.exists():
        raise FileNotFoundError(f"Missing valid_countries.json at {valid_path}")

    with open(valid_path, encoding="utf-8") as f:
        valid_list = json.load(f)

    valid_lookup = {v.lower().strip(): v for v in valid_list if isinstance(v, str)}

    alias_path = assets_dir / "aliases.json"
    aliases = {}
    if alias_path.exists():
        with open(alias_path, encoding="utf-8") as f:
            raw_aliases = json.load(f)
        aliases = {k.lower().strip(): v for k, v in raw_aliases.items()}

    return set(valid_lookup.keys()), valid_lookup, aliases


def extract_first_country(raw: str) -> str:
    """
    For multi-country strings (comma/slash-separated), extract the first country.
    Examples:
        "USA, Canada, Mexico"        → "USA"
        "Asia, Central America"      → "Asia"
        "Nicaragua, Colombia, Peru," → "Nicaragua"
        "USA"                        → "USA"
    """
    if not raw or not isinstance(raw, str):
        return ""
    parts = re.split(r"[,/;]", raw.strip())
    return parts[0].strip().rstrip(",. ")


def match_country(raw_value: str, valid_lookup: dict, aliases: dict) -> tuple[str, str]:
    """
    Map a single raw country string to a valid PlanetFWD COO value.

    Priority order:
        1. Null / empty           → "Unknown" (null)
        2. Regional / continental → "Unknown" (regional)
        3. Exact match            → valid name (exact)
        4. Alias lookup           → valid name (alias)
        5. Normalised match       → valid name (normalised)
        6. No match               → "Unknown" (no_match)

    Returns (validated_coo, match_method).
    """
    if not raw_value or (isinstance(raw_value, float) and pd.isna(raw_value)):
        return "Unknown", "null"

    normalized = raw_value.strip().lower()
    if not normalized:
        return "Unknown", "null"

    if normalized in REGIONAL_TERMS:
        return "Unknown", "regional"

    if normalized in valid_lookup:
        return valid_lookup[normalized], "exact"

    if normalized in aliases:
        return aliases[normalized], "alias"

    stripped = re.sub(
        r"^(the |republic of |islamic republic of |democratic |people's |federated states of )",
        "", normalized
    ).strip()

    if stripped in valid_lookup:
        return valid_lookup[stripped], "normalised"
    if stripped in aliases:
        return aliases[stripped], "alias"

    return "Unknown", "no_match"


def detect_coo_column(df: pd.DataFrame) -> str:
    """Auto-detect the column most likely to contain raw COO data."""
    candidates = [
        col for col in df.columns
        if any(kw in col.lower() for kw in ["country", "coo", "origin", "provenance"])
    ]
    if len(candidates) == 1:
        return candidates[0]
    if len(candidates) > 1:
        print(f"Multiple COO-like columns found: {candidates}")
        print(f"Defaulting to: '{candidates[0]}'. Use --col to override.")
        return candidates[0]
    raise ValueError(
        "Could not auto-detect COO column. "
        "Please specify it with --col <column_name>. "
        f"Available columns: {list(df.columns)}"
    )


def apply_styles(ws, out_df, mode="mapping"):
    """Apply header styling and colour-coding to the output worksheet."""
    from openpyxl.styles import Font, PatternFill, Alignment

    green_fill  = PatternFill("solid", fgColor="C6EFCE")
    amber_fill  = PatternFill("solid", fgColor="FFEB9C")
    red_fill    = PatternFill("solid", fgColor="FFC7CE")
    blue_fill   = PatternFill("solid", fgColor="DDEBF7")
    grey_fill   = PatternFill("solid", fgColor="EFEFEF")

    # Auto-size columns
    for col_idx, column in enumerate(out_df.columns, 1):
        max_len = max(
            len(str(column)),
            out_df.iloc[:, col_idx - 1].astype(str).map(len).max()
        )
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = min(max_len + 4, 60)

    # Header row — dark blue for validated columns, mid-blue for original data columns
    validated_cols = {"First Country Parsed", "Validated COO (PlanetFWD)", "Match Method",
                      "Raw Country of Origin"}
    header_dark = PatternFill("solid", fgColor="1F4E79")
    header_mid  = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        col_name = out_df.columns[cell.column - 1]
        cell.fill = header_dark if col_name in validated_cols else header_mid
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Colour-code Validated COO column
    validated_col_idx = list(out_df.columns).index("Validated COO (PlanetFWD)") + 1
    for row_idx, val in enumerate(out_df["Validated COO (PlanetFWD)"], start=2):
        cell = ws.cell(row=row_idx, column=validated_col_idx)
        cell.fill = amber_fill if val == "Unknown" else green_fill

    # Colour-code Match Method column
    method_col_idx = list(out_df.columns).index("Match Method") + 1
    method_colors = {
        "exact":      green_fill,
        "alias":      blue_fill,
        "normalised": blue_fill,
        "regional":   amber_fill,
        "no_match":   red_fill,
        "null":       grey_fill,
    }
    for row_idx, method in enumerate(out_df["Match Method"], start=2):
        cell = ws.cell(row=row_idx, column=method_col_idx)
        if method in method_colors:
            cell.fill = method_colors[method]


def validate_coo(input_file: str, col: str = None, sheet: str = None,
                 output_file: str = None, skill_dir: str = None,
                 mode: str = "mapping") -> str:
    """
    Main validation function.

    mode="mapping"  → output unique raw COO values with their validated mapping (default)
    mode="full"     → output the full original dataset with validated columns appended
    """
    if skill_dir is None:
        skill_dir = str(Path(__file__).parent.parent)

    print("Loading PlanetFWD valid country list and aliases...")
    valid_set, valid_lookup, aliases = load_assets(skill_dir)
    print(f"  Loaded {len(valid_lookup)} valid COO entries, {len(aliases)} aliases.")

    # ── Read input ────────────────────────────────────
    input_path = Path(input_file)
    print(f"\nReading input file: {input_path.name}")

    if input_path.suffix.lower() in [".xlsx", ".xlsm", ".xls"]:
        df = pd.read_excel(input_file, sheet_name=sheet) if sheet else pd.read_excel(input_file)
    elif input_path.suffix.lower() == ".csv":
        df = pd.read_csv(input_file)
        # Excel-exported CSVs often have a blank leading row which pandas treats as the
        # header, making every column "Unnamed: N". Detect this and re-read with skiprows=1.
        if all(str(c).startswith("Unnamed:") for c in df.columns):
            print("  ⚠️  Blank header row detected — re-reading with skiprows=1.")
            df = pd.read_csv(input_file, skiprows=1)
    elif input_path.suffix.lower() == ".tsv":
        df = pd.read_csv(input_file, sep="\t")
        if all(str(c).startswith("Unnamed:") for c in df.columns):
            print("  ⚠️  Blank header row detected — re-reading with skiprows=1.")
            df = pd.read_csv(input_file, sep="\t", skiprows=1)
    else:
        raise ValueError(f"Unsupported file type: {input_path.suffix}. Use .xlsx, .csv, or .tsv.")

    print(f"  Loaded {len(df):,} rows.")

    # ── Identify COO column ───────────────────────────
    if col is None:
        col = detect_coo_column(df)
    elif col not in df.columns:
        raise ValueError(f"Column '{col}' not found. Available: {list(df.columns)}")
    print(f"  Using column: '{col}'")

    # ── Build per-unique-value mapping ────────────────
    unique_raws = df[col].unique().tolist()
    print(f"  Found {len(unique_raws)} unique raw COO values.")

    print("\nRunning validation...")
    mapping = {}  # raw_value → (first_country, validated_coo, match_method)

    for raw in unique_raws:
        is_null = raw is None or (isinstance(raw, float) and pd.isna(raw))
        raw_str = "" if is_null else str(raw).strip()
        first_country = extract_first_country(raw_str)
        validated, method = match_country(first_country, valid_lookup, aliases)
        mapping[raw] = (first_country if first_country else None, validated, method)

    # ── Build output ──────────────────────────────────
    if mode == "full":
        # Append three new columns to the original dataframe
        out_df = df.copy()
        out_df["First Country Parsed"]    = df[col].map(lambda r: mapping[r][0])
        out_df["Validated COO (PlanetFWD)"] = df[col].map(lambda r: mapping[r][1])
        out_df["Match Method"]            = df[col].map(lambda r: mapping[r][2])
        sheet_name = "Full Data + COO Validation"

    else:  # mapping (default)
        rows = []
        for raw, (first, validated, method) in mapping.items():
            is_null = raw is None or (isinstance(raw, float) and pd.isna(raw))
            rows.append({
                "Raw Country of Origin":     None if is_null else raw,
                "First Country Parsed":      first,
                "Validated COO (PlanetFWD)": validated,
                "Match Method":              method,
            })

        out_df = pd.DataFrame(rows)
        # Sort: matched first, then Unknown; alphabetically within each group
        out_df["_sort"] = out_df["Validated COO (PlanetFWD)"].apply(
            lambda x: (1 if x == "Unknown" else 0, x or "")
        )
        out_df = out_df.sort_values("_sort").drop(columns=["_sort"]).reset_index(drop=True)
        sheet_name = "COO Validation"

    # ── Write output ──────────────────────────────────
    if output_file is None:
        suffix = "_validated_full" if mode == "full" else "_validated"
        output_file = str(input_path.parent / f"{input_path.stem}{suffix}.xlsx")

    print(f"\nWriting output ({mode} mode) to: {output_file}")

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        out_df.to_excel(writer, sheet_name=sheet_name, index=False)
        apply_styles(writer.sheets[sheet_name], out_df, mode=mode)

    # ── Print summary ─────────────────────────────────
    validated_series = out_df["Validated COO (PlanetFWD)"]
    total   = len(mapping)          # always based on unique values
    matched = sum(1 for _, v, _ in mapping.values() if v != "Unknown")
    unknown = total - matched

    print("\n" + "=" * 55)
    print(f"VALIDATION SUMMARY  [{mode.upper()} mode]")
    print("=" * 55)
    print(f"  Total unique raw values : {total:>6,}")
    print(f"  Successfully matched    : {matched:>6,} ({matched/total*100:.1f}%)")
    print(f"  Mapped to 'Unknown'     : {unknown:>6,} ({unknown/total*100:.1f}%)")
    if mode == "full":
        print(f"  Total rows in output    : {len(out_df):>6,}")
    print()

    methods = {}
    for _, _, m in mapping.values():
        methods[m] = methods.get(m, 0) + 1
    print("  Match method breakdown:")
    for method, count in sorted(methods.items(), key=lambda x: -x[1]):
        print(f"    {method:<20}: {count:>5}")

    if unknown > 0:
        print("\n  Values mapped to 'Unknown':")
        for raw, (_, v, _) in mapping.items():
            if v == "Unknown":
                print(f"    • {raw}")

    print("=" * 55)
    print(f"\nOutput saved: {output_file}")
    return output_file


def main():
    parser = argparse.ArgumentParser(
        description="PlanetFWD Country of Origin Validator",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__
    )
    parser.add_argument("input_file", help="Path to input file (.xlsx, .csv, or .tsv)")
    parser.add_argument("--col",   help="Column name containing raw COO data (auto-detected if omitted)")
    parser.add_argument("--sheet", help="Sheet name for Excel input (default: first sheet)")
    parser.add_argument("--output", help="Output file path")
    parser.add_argument("--skill-dir", default=str(Path(__file__).parent.parent),
                        help="Path to country_validation skill directory")
    parser.add_argument("--mode", choices=["mapping", "full"], default="mapping",
                        help="Output mode: 'mapping' (unique COO table, default) or "
                             "'full' (all original columns + validated columns appended)")

    args = parser.parse_args()

    try:
        output_path = validate_coo(
            input_file=args.input_file,
            col=args.col,
            sheet=args.sheet,
            output_file=args.output,
            skill_dir=args.skill_dir,
            mode=args.mode,
        )
        print(f"\n✅ Done! Output: {output_path}")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
